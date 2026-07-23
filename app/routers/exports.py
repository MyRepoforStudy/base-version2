from io import BytesIO
from typing import Iterable

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.session import get_db
from app.models import Host, HostFilesystem
from app.routers.common import (
    apply_host_filters,
    apply_operational_filters,
    format_uptime,
    health_for_host,
    last_reboot_at_for_host,
    lifecycle_status_for_host,
)
from app.services.zabbix_refresh import maybe_refresh_zabbix_cache

router = APIRouter(prefix="/exports", tags=["exports"])


def apply_sheet_style(ws, headers: Iterable[str]) -> None:
    header_fill = PatternFill("solid", fgColor="E9EEF5")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = max(14, len(header) + 2)


@router.get("/hosts.xlsx")
def export_hosts(
    environment: str | None = None,
    virtual: str | None = None,
    proxmox: str | None = None,
    system: str | None = None,
    criticality: str | None = None,
    lifecycle: str | None = None,
    health: str | None = None,
    db: Session = Depends(get_db),
):
    maybe_refresh_zabbix_cache(db)
    stmt = select(Host).order_by(Host.hostname)
    stmt = apply_host_filters(stmt, environment, virtual, proxmox, system, criticality)
    hosts = db.scalars(stmt).all()
    hosts = apply_operational_filters(hosts, lifecycle=lifecycle, health=health)

    headers = [
        "hostname",
        "ip_address",
        "environment",
        "datacenter",
        "virtual",
        "vendor",
        "model",
        "cpu_cores",
        "ram_gb",
        "uptime_seconds",
        "uptime",
        "last_reboot_at",
        "health_score",
        "health_status",
        "cpu_utilization_pct",
        "memory_utilization_pct",
        "load_average_1m",
        "root_disk_total_gb",
        "root_disk_used_gb",
        "root_disk_used_pct",
        "disk_max_mount",
        "disk_max_used_pct",
        "metrics_collected_at",
        "os_name",
        "os_family",
        "os_version",
        "os_support_end_date",
        "os_lifecycle_status",
        "owner",
        "department",
        "business_service",
        "criticality",
        "monitoring_status",
        "problem_count",
        "support_end_date",
        "zabbix_hostid",
        "zabbix_host_name",
        "zabbix_last_sync_at",
    ]
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Servers"
    ws.append(headers)
    for host in hosts:
        last_reboot = last_reboot_at_for_host(host)
        health = health_for_host(host)
        ws.append(
            [
                host.hostname,
                host.ip_address,
                host.environment,
                host.datacenter,
                "Virtual" if host.virtual else "Physical",
                host.vendor,
                host.model,
                host.cpu_cores,
                host.ram_gb,
                host.uptime_seconds,
                format_uptime(host.uptime_seconds),
                last_reboot.replace(tzinfo=None) if last_reboot else None,
                health.score,
                health.status,
                host.cpu_utilization_pct,
                host.memory_utilization_pct,
                host.load_average_1m,
                host.root_disk_total_gb,
                host.root_disk_used_gb,
                host.root_disk_used_pct,
                host.disk_max_mount,
                host.disk_max_used_pct,
                host.metrics_collected_at.replace(tzinfo=None) if host.metrics_collected_at else None,
                host.os_name,
                host.os_family,
                host.os_version,
                host.os_support_end_date,
                lifecycle_status_for_host(host),
                host.owner,
                host.department,
                host.business_service,
                host.criticality,
                host.monitoring_status,
                host.problem_count,
                host.support_end_date,
                host.zabbix_hostid,
                host.zabbix_host_name,
                host.zabbix_last_sync_at.replace(tzinfo=None) if host.zabbix_last_sync_at else None,
            ]
        )
    apply_sheet_style(ws, headers)

    filesystem_headers = [
        "hostname",
        "mount_point",
        "total_gb",
        "used_gb",
        "used_pct",
        "collected_at",
    ]
    filesystem_ws = workbook.create_sheet("Filesystems")
    filesystem_ws.append(filesystem_headers)
    hostnames_by_id = {host.id: host.hostname for host in hosts}
    if hostnames_by_id:
        filesystems = db.scalars(
            select(HostFilesystem)
            .where(HostFilesystem.host_id.in_(list(hostnames_by_id)))
            .order_by(HostFilesystem.host_id, HostFilesystem.mount_point)
        ).all()
        for filesystem in filesystems:
            filesystem_ws.append(
                [
                    hostnames_by_id.get(filesystem.host_id),
                    filesystem.mount_point,
                    filesystem.total_gb,
                    filesystem.used_gb,
                    filesystem.used_pct,
                    filesystem.collected_at.replace(tzinfo=None),
                ]
            )
    apply_sheet_style(filesystem_ws, filesystem_headers)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="server_inventory_hosts.xlsx"'},
    )
